"""
File Output Utilities.
Generates final Excel output with main data sheet and error sheet.
"""

import pandas as pd
from pathlib import Path
from typing import List, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from ..schemas.validation_report import ValidationReport, ValidationError
from ..config import get_settings


class FileOutput:
    """
    Generates formatted Excel output files.
    Creates a main data sheet and an error sheet for failed rows.
    """
    
    # Style constants
    HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF")
    ERROR_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    WARNING_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    THIN_BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    def __init__(self, output_path: Optional[str] = None):
        """
        Initialize file output handler.
        
        Args:
            output_path: Path for output file. If None, uses settings.output_dir
        """
        self.settings = get_settings()
        self.output_path = Path(output_path) if output_path else None
        self._workbook: Optional[Workbook] = None
    
    def generate(
        self,
        df: pd.DataFrame,
        validation_report: Optional[ValidationReport] = None,
        output_filename: str = "transformed_output.xlsx",
        include_error_sheet: bool = True,
    ) -> Path:
        """
        Generate the output Excel file.
        
        Args:
            df: Transformed DataFrame
            validation_report: Optional validation report with errors
            output_filename: Name of output file
            include_error_sheet: Whether to create an error sheet
            
        Returns:
            Path to the generated file
        """
        # Ensure output directory exists
        self.settings.ensure_directories()
        
        output_path = self.output_path or (self.settings.output_dir / output_filename)
        
        # Create workbook
        self._workbook = Workbook()
        
        # Create main data sheet
        self._create_data_sheet(df, "Transformed Data")
        
        # Create error sheet if there are errors
        if include_error_sheet and validation_report and validation_report.errors:
            self._create_error_sheet(df, validation_report.errors)
        
        # Create summary sheet
        if validation_report:
            self._create_summary_sheet(validation_report)
        
        # Remove default empty sheet if we created others
        if "Sheet" in self._workbook.sheetnames and len(self._workbook.sheetnames) > 1:
            del self._workbook["Sheet"]
        
        # Save workbook
        self._workbook.save(output_path)
        
        return output_path
    
    def _create_data_sheet(
        self,
        df: pd.DataFrame,
        sheet_name: str = "Data"
    ):
        """Create the main data sheet."""
        if "Data" in self._workbook.sheetnames:
            ws = self._workbook["Data"]
        else:
            ws = self._workbook.create_sheet(sheet_name)
        
        # Write DataFrame
        for row_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = self.THIN_BORDER
                
                # Style header row
                if row_idx == 1:
                    cell.fill = self.HEADER_FILL
                    cell.font = self.HEADER_FONT
                    cell.alignment = Alignment(horizontal='center')
        
        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
    
    def _create_error_sheet(
        self,
        df: pd.DataFrame,
        errors: List[ValidationError]
    ):
        """Create sheet showing rows with errors."""
        ws = self._workbook.create_sheet("Errors")
        
        # Get unique error row indices
        error_rows = set(e.row_index for e in errors)
        
        # Header row
        headers = ["Row #", "Error Type", "Column", "Value", "Issue", "Suggested Fix"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.border = self.THIN_BORDER
        
        # Error rows
        for row_idx, error in enumerate(errors, 2):
            row_data = [
                error.row_index + 1,  # 1-indexed for user display
                error.severity.upper(),
                error.column,
                error.value or "",
                error.issue,
                error.suggested_fix or ""
            ]
            
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = self.THIN_BORDER
                
                # Color based on severity
                if error.severity == "error":
                    cell.fill = self.ERROR_FILL
                else:
                    cell.fill = self.WARNING_FILL
        
        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
    
    def _create_summary_sheet(self, validation_report: ValidationReport):
        """Create a summary sheet with validation statistics."""
        ws = self._workbook.create_sheet("Summary")
        
        # Title
        ws.cell(row=1, column=1, value="Transformation Summary")
        ws.cell(row=1, column=1).font = Font(bold=True, size=14)
        
        # Statistics
        stats = [
            ("Status", validation_report.status.upper()),
            ("Total Rows", validation_report.total_rows),
            ("Successful Rows", validation_report.successful_rows),
            ("Failed Rows", validation_report.failed_rows),
            ("Warning Rows", validation_report.warning_rows),
            ("Quality Score", f"{validation_report.quality_score:.1f}%"),
        ]
        
        for row_idx, (label, value) in enumerate(stats, 3):
            ws.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
            ws.cell(row=row_idx, column=2, value=value)
        
        # Column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 30


def save_output(
    df: pd.DataFrame,
    validation_report: Optional[ValidationReport] = None,
    output_path: Optional[str] = None,
    filename: str = "output.xlsx"
) -> Path:
    """
    Convenience function to save output.
    
    Args:
        df: Transformed DataFrame
        validation_report: Optional validation report
        output_path: Custom output path
        filename: Output filename
        
    Returns:
        Path to saved file
    """
    output = FileOutput(output_path)
    return output.generate(df, validation_report, filename)
